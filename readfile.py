import openpyxl

##Open and read excel file pretty much self explanatory.
class Readfile:
    def __init__(self, filename):
        self.wb_obj = openpyxl.load_workbook(filename)
        self.sheet_obj = self.wb_obj.active

    def get_maxrow(self):
        max_row = 0
        for row in self.sheet_obj:
            if not all([cell.value == None for cell in row]):
                max_row += 1
        return max_row

    def get_value(self, row, col):
        return self.sheet_obj.cell(row, col)

    def set_value(self, row, col, val):
        tmp = self.sheet_obj.cell(row,col)
        tmp.value = val

    def set_list(self, row, col, lst):
        for i in lst:
            tmp = self.sheet_obj.cell(row,col)
            tmp.value = i
            col += 1

    def save(self, file):
        self.wb_obj.save(file)