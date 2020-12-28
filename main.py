import mylib
import openpyxl

path = "newform.xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

for i in range(2, max_row + 1):
    OGR_NUM = sheet_obj.cell(row = i, column = 5) 
    TC_KIM = sheet_obj.cell(row = i, column = 6)
    c1 = sheet_obj.cell(row = i, column = 7)
    c2 = sheet_obj.cell(row = i, column = 8)
    c3 = sheet_obj.cell(row = i, column = 9)
    array = mylib.getnum(OGR_NUM.value, TC_KIM.value)
    c1.value = array[0]
    c2.value = array[1]
    c3.value = array[2]
    print(array)
    wb_obj.save("newform2.xlsx")

